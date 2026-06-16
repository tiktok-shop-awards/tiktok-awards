import json
import os
import re
from pypinyin import pinyin, Style
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font

# 工作目录
WORK_DIR = '/app/data/所有对话/主对话/'
DATA_DIR = '/app/data/所有对话/主对话/new-version-dev/employee_data/'

# 1. 加载已有映射 (约240条)
with open(WORK_DIR + 'cjk_name_mapping_from_tables.json', 'r', encoding='utf-8') as f:
    table_mapping = json.load(f)
print(f"Loaded {len(table_mapping)} mappings from table")

# 2. 从飞书员工表构建邮箱前缀→英文名映射
email_to_english = {}

# 加载所有页面数据
page_files = [f for f in os.listdir(DATA_DIR) if f.startswith('page_') and f.endswith('.json')]
page_files.sort(key=lambda x: int(x.replace('page_', '').replace('.json', '')))

all_employees = []
for pf in page_files:
    with open(DATA_DIR + pf, 'r', encoding='utf-8') as f:
        data = json.load(f)
        records = data.get('data', {}).get('data', [])
        all_employees.extend(records)

print(f"Loaded {len(all_employees)} employee records from {len(page_files)} pages")

# 构建邮箱→英文名映射
for name, email in all_employees:
    if not email or not name:
        continue
    # 提取邮箱前缀
    prefix = email.split('@')[0]
    # 只保留没有数字后缀的前缀（标准格式）
    clean_prefix = re.sub(r'\.\d+$', '', prefix)
    # 存储（保留原始顺序，最后一个为准）
    if clean_prefix and '.' not in clean_prefix and len(clean_prefix) > 2:
        email_to_english[clean_prefix.lower()] = name

print(f"Built email-to-english mapping with {len(email_to_english)} entries")

# 3. 从员工表构建中文名→英文名映射
chinese_name_mapping = {}
for name, email in all_employees:
    if not name:
        continue
    # 提取纯中文名（如果有英文名在后面，如"王杰 Benjamin"）
    parts = re.split(r'[\s,]+', name)
    chinese_part = None
    english_part = None
    
    for part in parts:
        if re.search(r'[\u4e00-\u9fff]', part):
            chinese_part = part
        elif re.match(r'^[A-Za-z]+$', part) or '.' in part:
            english_part = part
    
    if chinese_part and english_part:
        chinese_name_mapping[chinese_part] = english_part

print(f"Built Chinese name mapping with {len(chinese_name_mapping)} entries")

# 4. 定义翻译函数
def extract_chinese_name(full_name):
    """从完整名称中提取中文名"""
    if not full_name:
        return None
    # 匹配中文
    match = re.search(r'[\u4e00-\u9fff]+', full_name)
    if match:
        return match.group()
    return None

def get_pinyin(chinese_name):
    """获取中文名的拼音"""
    try:
        pys = [p[0] for p in pinyin(chinese_name, style=Style.NORMAL)]
        return ''.join(pys)
    except:
        return None

def translate_name(full_name, email):
    """翻译中文名为英文名"""
    result = {
        'chinese_name': None,
        'english_name': None,
        'source': None,
        'confidence': None
    }
    
    # 提取中文名
    chinese = extract_chinese_name(full_name)
    if not chinese:
        result['english_name'] = full_name  # 没有中文，保持原样
        result['source'] = 'No Chinese'
        result['confidence'] = 'N/A'
        return result
    
    result['chinese_name'] = chinese
    
    # 优先级1: 直接从table映射
    if chinese in table_mapping:
        result['english_name'] = table_mapping[chinese]
        result['source'] = 'Table'
        result['confidence'] = 'High'
        return result
    
    # 优先级2: 从员工表中文名映射
    if chinese in chinese_name_mapping:
        result['english_name'] = chinese_name_mapping[chinese]
        result['source'] = 'Employee Table'
        result['confidence'] = 'High'
        return result
    
    # 优先级3: 从邮箱匹配
    if email:
        email_prefix = email.split('@')[0].lower()
        # 尝试匹配带点的邮箱前缀
        clean_prefix = re.sub(r'\.\d+$', '', email_prefix)
        # 提取英文名部分（如果有）
        if '.' in clean_prefix:
            parts = clean_prefix.split('.')
            for part in parts:
                if len(part) > 2 and part not in ['bytedance', 'com']:
                    # 检查是否是英文名
                    if re.match(r'^[a-z]+$', part):
                        result['english_name'] = part.capitalize()
                        result['source'] = 'Email'
                        result['confidence'] = 'Medium'
                        return result
        
        # 尝试匹配纯前缀
        if clean_prefix in email_to_english:
            result['english_name'] = email_to_english[clean_prefix]
            result['source'] = 'Email Lookup'
            result['confidence'] = 'Medium'
            return result
    
    # 优先级4: 从姓名中提取英文部分
    parts = re.split(r'[\s,()（）]+', full_name)
    for part in parts:
        if re.match(r'^[A-Z][a-z]+$', part) and len(part) > 2:
            # 可能是英文名
            result['english_name'] = part
            result['source'] = 'Extracted'
            result['confidence'] = 'Low'
            return result
    
    # 优先级5: 使用拼音
    pinyin_name = get_pinyin(chinese)
    if pinyin_name:
        result['english_name'] = pinyin_name.capitalize()
        result['source'] = 'Pinyin'
        result['confidence'] = 'Low'
        return result
    
    result['source'] = 'Not Found'
    result['confidence'] = 'None'
    return result

# 5. 读取源文件并翻译
wb = openpyxl.load_workbook(WORK_DIR + 'CJK_Name_Translation_v2.xlsx')
ws = wb.active

# 添加新列
headers = [cell.value for cell in ws[1]]
headers.extend(['Translated English Name', 'Translation Source', 'Confidence'])
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 翻译每一行
stats = {'high': 0, 'medium': 0, 'low': 0, 'none': 0, 'no_chinese': 0}
for row_idx in range(2, ws.max_row + 1):
    full_name = ws.cell(row=row_idx, column=2).value  # Chinese Name
    email = ws.cell(row=row_idx, column=5).value  # Email(s)
    
    # 如果已有英文名，保留
    existing_english = ws.cell(row=row_idx, column=3).value
    if existing_english:
        ws.cell(row=row_idx, column=7, value=existing_english)
        ws.cell(row=row_idx, column=8, value='Existing')
        ws.cell(row=row_idx, column=9, value='High')
        stats['high'] += 1
        continue
    
    result = translate_name(full_name, email)
    
    ws.cell(row=row_idx, column=7, value=result['english_name'])
    ws.cell(row=row_idx, column=8, value=result['source'])
    ws.cell(row=row_idx, column=9, value=result['confidence'])
    
    # 统计
    if result['confidence'] == 'High':
        stats['high'] += 1
    elif result['confidence'] == 'Medium':
        stats['medium'] += 1
    elif result['confidence'] == 'Low':
        stats['low'] += 1
    elif result['confidence'] == 'None':
        stats['none'] += 1
    else:
        stats['no_chinese'] += 1

print(f"\nTranslation Statistics:")
print(f"  High confidence: {stats['high']}")
print(f"  Medium confidence: {stats['medium']}")
print(f"  Low confidence: {stats['low']}")
print(f"  Not found: {stats['none']}")
print(f"  No Chinese: {stats['no_chinese']}")

# 6. 保存文件
output_path = WORK_DIR + 'CJK_Name_Translation_v3.xlsx'
wb.save(output_path)
print(f"\nSaved to: {output_path}")
